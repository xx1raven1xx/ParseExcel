import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime
import time
import sys, winsound

sound_path = "D:\\python\\ParseEXEL\\beep.wav"

start_time = time.time()

print('Подготовьте файл с именем и расширением data.xlsx.')
input('Далее нажмите ENTER.')

#при встрече в ячейке времени значения 00:00:00 выдавал ошибку по типу данных 'TipeError'

FillR = PatternFill(start_color='00fcd4d1',
                   end_color='FFFF0000',
                   fill_type='solid')
FillY = PatternFill(start_color='00fffbcc',
                   end_color='FFFF0000',
                   fill_type='solid')
Color1 = True


A0 = datetime.time(0, 00, 00)
A8 = datetime.time(8, 00, 00)
A830 = datetime.time(8, 30, 00)
A12 = datetime.time(12, 00, 00)
A13 = datetime.time(13, 00, 00)
A16 = datetime.time(16, 00, 00)
A17 = datetime.time(17, 00, 00)
A1705 = datetime.time(17, 5, 00)
A19 = datetime.time(19, 00, 00)
A20 = datetime.time(20, 00, 00)
A2005 = datetime.time(20, 5, 00)
A2030 = datetime.time(20, 30, 00)
A24 = datetime.time(23, 59, 59)
min20 = datetime.time(00, 20, 00)


Date_in_sheet = 0
Time_in_sheet = 2
KPP = 3
Work_post = 4
InOut = 5
FIO = 8
Destination = 9


#Сохраняет в список нужные нам данные и потом сам список сохраняется в файл
def sv(ws, dt, tm, kpp, pst, nm, dst, o, reserv='', Paint = False):
#активный лист, дата, время, кпп, должность, имя, подразделение, порядковый номер, резерв
    global Color1
    a = o + 1
    o = str(o)
    a = str(a)
    ca = 'A' + o
    cb = 'B' + o
    cc = 'C' + o
    cd = 'D' + o
    cd1 = 'D' + a
    ce = 'E' + o
    cf = 'F' + o
    cg = 'G' + o
    ws[ca] = dt
    ws[cb] = tm
    ws[cc] = kpp
    ws[cd] = pst
    ws[ce] = nm
    ws[cf] = dst
    ws[cg] = reserv
    o = int(o)
    if Paint:
        if pst == 'ВЫХОД':
            if Color1:
                ws[cd].fill = FillR
                ws[cd1].fill = FillR
                Color1 = False
            else:
                ws[cd].fill = FillY
                ws[cd1].fill = FillY
                Color1 = True

#возвращает самую первую отметку 'ВХОД' или 'ВЫХОД' в этот день
def ft(name, date, start=1):
    #нужно считать день, время и смотреть самую раннюю отметку в данный день. Возвращает ВХОД или ВЫХОД.
    global cell_range
    a = A24
    for i in range(start, ws.max_row-1):
        if cell_range[i][FIO].value == name:
            if cell_range[i][0].value == date:
                if a > cell_range[i][Time_in_sheet].value:
                    a = cell_range[i][Time_in_sheet].value
                    b = cell_range[i][InOut].value
                    
        else:
            break
            
    return b

#возвращает последнюю отметку 'ВХОД' или 'ВЫХОД' в этот день
def et(name, date, start=1):
    #нужно считать день, время и смотреть самую позднюю отметку в данный день. Возвращает ВХОД или ВЫХОД.
    global cell_range
    a = A0
    for i in range(start, ws.max_row-1):
        if cell_range[i][FIO].value == name:
            if cell_range[i][0].value == date:
                if a < cell_range[i][Time_in_sheet].value:
                    a = cell_range[i][Time_in_sheet].value
                    b = cell_range[i][InOut].value
                    
        else:
            break
            
    return b

#выводит минимальное время входа в определенный день
def minT(name, date, start=1):
    global cell_range
    a = A24
    for i in range(start, ws.max_row-1):
        if cell_range[i][FIO].value == name:
            if cell_range[i][0].value == date:
                if cell_range[i][InOut].value == 'ВХОД':
                    if a > cell_range[i][Time_in_sheet].value:
                        a = cell_range[i][Time_in_sheet].value
        else:
            break
    return a

#выводит максимальное время выхода в определенный день
def maxT(name, date, start=1):
    global cell_range
    a = A0
    for i in range(start, ws.max_row-1):
        if cell_range[i][FIO].value == name:
            if cell_range[i][0].value == date:
                if cell_range[i][InOut].value == 'ВЫХОД':
                    if a < cell_range[i][Time_in_sheet].value:
                        a = cell_range[i][Time_in_sheet].value
        else:
            break
    return a



try:
    wb = openpyxl.load_workbook(filename = 'data.xlsx', read_only=True)
except:
    print('Проверте файл!!!')
    input('Нажмите ENTER.')
    sys.exit()

#sheet = wb['Лист1']                     #по сути одно и то же что и 
ws = wb.active                          #это

print('Начинаем обработку данных...')

wb1 = Workbook()
wsss = wb1.active 
wb1.remove(wsss)                        #удаляем активный лист - он нам не нужен.

wb1.create_sheet('ВХ0800')
ws1 = wb1['ВХ0800']

wb1.create_sheet('ВХдо2000')
ws2 = wb1['ВХдо2000']

wb1.create_sheet('ВЫ1700')
ws3 = wb1['ВЫ1700']

wb1.create_sheet('ВЫ2000')
ws4 = wb1['ВЫ2000']

wb1.create_sheet('До1705')
ws5 = wb1['До1705']

wb1.create_sheet('До2005')
ws6 = wb1['До2005']

wb1.create_sheet('20_MIN')
ws7 = wb1['20_MIN']


#c = 'F2500'
#colC = ws['C']                          #создает кортеж что ли???
#row7 = ws[7]
cell_range = ws['a1':'j'+str(ws.max_row-1)]
#сам кортеж не изменяемый, но вот списки в нем... Или это не списки а обьекты?
print('Файл имеет ', ws.max_row-1, ' строк \n')

l = 1
k = 1
j = 1
h = 1
j1 = 1
h1 = 1
f = 1

for i in range(2,ws.max_row-1):
    start1 = time.time()
    i1 = i - 1
    endtime = datetime.datetime.combine(datetime.datetime.strptime(cell_range[i1+1][Date_in_sheet].value, '%Y.%m.%d').date(), cell_range[i1+1][Time_in_sheet].value)
    starttime = datetime.datetime.combine(datetime.datetime.strptime(cell_range[i1][Date_in_sheet].value, '%Y.%m.%d').date(), cell_range[i1][Time_in_sheet].value)
    STED = endtime - starttime
    if cell_range[i1][Time_in_sheet].value < A12 or cell_range[i1+1][Time_in_sheet].value > A13:
        if cell_range[i1][InOut].value == 'ВЫХОД':
            if cell_range[i1+1][InOut].value == 'ВХОД':
                if cell_range[i1][Date_in_sheet].value == cell_range[i1+1][Date_in_sheet].value:
                    if STED > datetime.timedelta(minutes = 20):
                        sv(ws7, cell_range[i1][Date_in_sheet].value, cell_range[i1][Time_in_sheet].value,cell_range[i1][KPP].value, cell_range[i1][InOut].value, cell_range[i1][Work_post].value, cell_range[i1][FIO].value, f, Paint = True)
                        f += 1
                        sv(ws7, cell_range[i1+1][Date_in_sheet].value, cell_range[i1+1][Time_in_sheet].value, cell_range[i1+1][KPP].value,cell_range[i1+1][InOut].value, cell_range[i1+1][Work_post].value, cell_range[i1+1][FIO].value, f)
                        f += 1

    if cell_range[i][0].value == cell_range[i-1][0].value:          #если дата и предыдущая дата равны то пропустить один цикл???  накуя я это написал то???
        #print('test')
        continue

    cellDAY = cell_range[i][Date_in_sheet].value    # Дата на листе в ячейке
    cellKPP = cell_range[i][KPP].value              # КПП
    cellPST = cell_range[i][Work_post].value        # Должность
    cellFIO = cell_range[i][FIO].value              # ФИО
    cellDST = cell_range[i][Destination].value      # Подразделение
    fist_time = ft(cellFIO, cellDAY, i)             # Первая отметка
    min1 = minT(cellFIO, cellDAY, i)                # Минимальное время входа
    end_time = et(cellFIO, cellDAY, i)              # Последняя отметка
    max1 = maxT(cellFIO, cellDAY, i)                # Максимальное время выхода
    
    #поздний вход 08,00 - 08,30
    if fist_time == 'ВХОД':
        if min1 >= A8 and min1 <= A830:
            sv(ws1, cellDAY, min1, cellKPP, cellPST, cellFIO, cellDST, l)
            l += 1
        #поздний вход 20,00 - 20,30
        elif min1 >= A20 and min1 <= A2030:
            sv(ws2, cellDAY, min1, cellKPP, cellPST, cellFIO, cellDST, k)
            k += 1
    #ранний выход 16,00 - 17,00
    if end_time == 'ВЫХОД':
        if max1 <= A17 and max1 >= A16:
            sv(ws3, cellDAY, max1, cellKPP, cellPST, cellFIO, cellDST, j)
            j += 1
        #ранний выход 17,00 - 17,05
        elif max1 <= A1705 and max1 >= A17:
            sv(ws5, cellDAY, max1, cellKPP, cellPST, cellFIO, cellDST, j1)
            j1 += 1
        #ранний выход 19,00 - 20,00
        elif max1 <= A20 and max1 >= A19:
            sv(ws4, cellDAY, max1, cellKPP, cellPST, cellFIO, cellDST, h)
            h += 1
        #ранний выход 20,00 - 20,05
        elif max1 <= A2005 and max1 >= A20:
            sv(ws6, cellDAY, max1, cellKPP, cellPST, cellFIO, cellDST, h1)
            h1 += 1

    end1 = time.time()
    sys.stdout.write("\rКоличество обработанных строк %i" % i)
    sys.stdout.flush()


#перед сохранением ОБЯЗАТЕЛЬНО закрыть файл в Офисе.

#Я в душе не **у как сделать короче данный кусок кода. Может быть с помощью функции exec()
ws1.column_dimensions['A'].width = (10*2.3)/1.96
ws1.column_dimensions['C'].width = (10*3.6)/1.96
ws1.column_dimensions['D'].width = (10*10.4)/1.96
ws1.column_dimensions['E'].width = (10*6.9)/1.96
ws1.column_dimensions['F'].width = (10*7.9)/1.96

ws2.column_dimensions['A'].width = (10*2.3)/1.96
ws2.column_dimensions['C'].width = (10*3.6)/1.96
ws2.column_dimensions['D'].width = (10*10.4)/1.96
ws2.column_dimensions['E'].width = (10*6.9)/1.96
ws2.column_dimensions['F'].width = (10*7.9)/1.96

ws3.column_dimensions['A'].width = (10*2.3)/1.96
ws3.column_dimensions['C'].width = (10*3.6)/1.96
ws3.column_dimensions['D'].width = (10*10.4)/1.96
ws3.column_dimensions['E'].width = (10*6.9)/1.96
ws3.column_dimensions['F'].width = (10*7.9)/1.96

ws4.column_dimensions['A'].width = (10*2.3)/1.96
ws4.column_dimensions['C'].width = (10*3.6)/1.96
ws4.column_dimensions['D'].width = (10*10.4)/1.96
ws4.column_dimensions['E'].width = (10*6.9)/1.96
ws4.column_dimensions['F'].width = (10*7.9)/1.96

ws5.column_dimensions['A'].width = (10*2.3)/1.96
ws5.column_dimensions['C'].width = (10*3.6)/1.96
ws5.column_dimensions['D'].width = (10*10.4)/1.96
ws5.column_dimensions['E'].width = (10*6.9)/1.96
ws5.column_dimensions['F'].width = (10*7.9)/1.96

ws6.column_dimensions['A'].width = (10*2.3)/1.96
ws6.column_dimensions['C'].width = (10*3.6)/1.96
ws6.column_dimensions['D'].width = (10*10.4)/1.96
ws6.column_dimensions['E'].width = (10*6.9)/1.96
ws6.column_dimensions['F'].width = (10*7.9)/1.96

ws7.column_dimensions['A'].width = (10*2.3)/1.96
ws7.column_dimensions['C'].width = (10*2.9)/1.96
ws7.column_dimensions['D'].width = (10*1.7)/1.96
ws7.column_dimensions['E'].width = (10*7.9)/1.96
ws7.column_dimensions['F'].width = (10*6.8)/1.96
ws7.column_dimensions['G'].width = (10*1.7)/1.96


wb1.save('result.xlsx')
print("\n--- %s секунд ---" % (time.time() - start_time))
print('ВЫПОЛНЕНО.')
print('Файл result.xlsx содержит обработанные данные.')
winsound.PlaySound(sound_path, winsound.SND_FILENAME)
input('Обработка окончена. Нажмите ENTER.')
